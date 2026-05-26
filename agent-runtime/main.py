"""
Yuno AI Agent Runtime — FastAPI server
Provides SSE-streaming workflow execution and direct agent calls.
"""
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

import os
import sys
import json
import asyncio
import shutil
import subprocess
from urllib.parse import unquote
from io import BytesIO
from pathlib import Path
from typing import Any
from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from dotenv import load_dotenv

UPLOAD_DIR = Path(__file__).parent / "workspace" / "uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

from runtime.langgraph_runner import execute_langgraph_workflow
from runtime.agent_executor import run_agent
from runtime.orchestrator import run_autonomous, generate_agent_intro
from runtime.workflow_validator import validate_workflow_inputs

load_dotenv()

# ── LangSmith tracing (auto-enabled when LANGSMITH_API_KEY is present) ────────
if os.getenv("LANGSMITH_API_KEY", "").strip():
    os.environ.setdefault("LANGCHAIN_TRACING_V2", "true")
    os.environ.setdefault("LANGCHAIN_PROJECT",    "yuno-ai")
    os.environ.setdefault("LANGCHAIN_ENDPOINT",   "https://api.smith.langchain.com")

app = FastAPI(title="Yuno AI Runtime", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


class WorkflowRequest(BaseModel):
    run_id: str
    workflow: dict
    input: dict = {}            # structured: {message, pdf_path, recipient_emails, ...}
    agent_memories: dict = {}
    openai_api_key: str = ""
    google_credentials: dict | None = None
    calendar_timezone: str = "UTC"


class AgentRequest(BaseModel):
    agent: dict
    message: str
    chat_id: str | None = None
    openai_api_key: str = ""
    google_credentials: dict | None = None   # service account JSON passed from UI
    calendar_timezone: str = "UTC"


@app.get("/health")
async def health():
    return {"status": "ok", "service": "yuno-ai-runtime"}


# ── Token cost monitoring ─────────────────────────────────────────────────────

@app.get("/token-stats")
async def token_stats():
    """
    Aggregated token usage + estimated cost.

    Returns LangSmith data when LANGSMITH_API_KEY is configured,
    otherwise falls back to the local workspace/token_usage.json store.

    Response shape:
      {
        source: "langsmith" | "local",
        langsmith_project: str | null,
        today:  { runs, input_tokens, output_tokens, total_tokens, cost_usd },
        week:   { ... },
        month:  { ... },
        daily:  { "YYYY-MM-DD": { ... } },   # last 7 days
        models: { "gpt-4o": { ... } },        # last 30 days
      }
    """
    from runtime.token_tracker import get_stats
    return get_stats()


@app.delete("/token-stats")
async def clear_token_stats():
    """Wipe the local token_usage.json (useful for resetting the counter)."""
    from runtime.token_tracker import TRACKER_FILE
    try:
        if TRACKER_FILE.exists():
            TRACKER_FILE.unlink()
        return {"cleared": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class WorkflowValidateRequest(BaseModel):
    workflow:  dict          # same shape as WorkflowRequest.workflow
    collected: dict = {}     # all user-collected inputs so far


@app.post("/workflow/validate")
async def validate_workflow(req: WorkflowValidateRequest):
    """
    Validate collected workflow inputs against each agent's required tools.
    Returns { valid: bool, errors: [...] } before the workflow executes.
    Each error: { agent_name, tool, field, label, message }.
    """
    errors = validate_workflow_inputs(req.workflow, req.collected)
    return {"valid": len(errors) == 0, "errors": errors}


class AgentIntroRequest(BaseModel):
    agent: dict
    openai_api_key: str = ""


@app.post("/agent/intro")
async def agent_intro(req: AgentIntroRequest):
    """
    Orchestrator hook: called when an agent chat is opened.
    Reads the agent's system_prompt and tools[], uses the LLM to generate
    a dynamic context-aware greeting (instead of a hardcoded one).
    """
    api_key = req.openai_api_key or os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        # Return a safe fallback if no key is configured yet
        name = req.agent.get("name", "Agent")
        desc = req.agent.get("description") or req.agent.get("system_prompt", "")
        return {"greeting": f"Hi! I'm {name}. {desc[:120] or 'How can I help you?'}"}

    greeting = await generate_agent_intro(req.agent, api_key)
    return {"greeting": greeting}


@app.post("/execute/workflow")
async def execute_workflow(req: WorkflowRequest):
    """Execute a workflow and stream events as Server-Sent Events."""
    api_key = req.openai_api_key or os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="OPENAI_API_KEY not configured")

    event_queue: asyncio.Queue = asyncio.Queue()

    async def event_callback(event_type: str, data: dict):
        event = {"type": event_type, **data}
        await event_queue.put(event)

    async def run_workflow():
        try:
            result = await execute_langgraph_workflow(
                workflow=req.workflow,
                initial_input=req.input,
                api_key=api_key,
                agent_memories=req.agent_memories,
                event_callback=event_callback,
                google_credentials=req.google_credentials,
                calendar_timezone=req.calendar_timezone,
            )
            await event_queue.put({
                "type":         "workflow_complete",
                "run_id":       req.run_id,
                "output":       result.get("agent_outputs", {}),
                "final_output": result.get("final_output", ""),
                "conversation": result.get("conversation", []),
                "token_usage":  result.get("token_usage", {}),
            })
        except Exception as e:
            await event_queue.put({
                "type": "workflow_error",
                "run_id": req.run_id,
                "error": str(e)
            })
        finally:
            await event_queue.put(None)  # Sentinel to end stream

    async def event_generator():
        task = asyncio.create_task(run_workflow())
        while True:
            try:
                event = await asyncio.wait_for(event_queue.get(), timeout=300)
                if event is None:
                    break
                yield f"data: {json.dumps(event)}\n\n"
            except asyncio.TimeoutError:
                yield f"data: {json.dumps({'type': 'keepalive'})}\n\n"
        await task

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no"
        }
    )


@app.post("/execute/agent")
async def execute_agent(req: AgentRequest):
    """Execute a single agent and return the result directly."""
    api_key = req.openai_api_key or os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="OPENAI_API_KEY not configured")

    messages = [{"role": "user", "content": req.message}]

    try:
        result = await run_agent(
            agent_config=req.agent,
            messages=messages,
            api_key=api_key,
            memory=None,
            event_callback=None,
            google_credentials=req.google_credentials,
            calendar_timezone=req.calendar_timezone,
        )
        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/execute/agent/stream")
async def execute_agent_stream(req: AgentRequest):
    """Execute a single agent and stream tool calls, results, and final output as SSE."""
    api_key = req.openai_api_key or os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="OPENAI_API_KEY not configured")

    messages = [{"role": "user", "content": req.message}]
    event_queue: asyncio.Queue = asyncio.Queue()

    async def event_callback(event_type: str, data: dict):
        await event_queue.put({"type": event_type, **data})

    async def run():
        try:
            result = await run_agent(
                agent_config=req.agent,
                messages=messages,
                api_key=api_key,
                memory=None,
                event_callback=event_callback,
                google_credentials=req.google_credentials,
                calendar_timezone=req.calendar_timezone,
            )
            await event_queue.put({
                "type":        "agent_complete",
                "output":      result.get("output", ""),
                "token_usage": result.get("token_usage", {}),
            })
        except Exception as e:
            await event_queue.put({"type": "agent_error", "error": str(e)})
        finally:
            await event_queue.put(None)

    async def event_generator():
        task = asyncio.create_task(run())
        while True:
            try:
                event = await asyncio.wait_for(event_queue.get(), timeout=120)
                if event is None:
                    break
                yield f"data: {json.dumps(event)}\n\n"
            except asyncio.TimeoutError:
                yield f"data: {json.dumps({'type': 'keepalive'})}\n\n"
        await task

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


class AutonomousRequest(BaseModel):
    goal: str
    agents: list[dict] = []          # registered agents from the DB
    openai_api_key: str = ""


@app.post("/execute/autonomous")
async def execute_autonomous(req: AutonomousRequest):
    """
    Fully autonomous execution — the LLM orchestrator plans and executes everything.
    No workflow needed. Just give it a goal.
    """
    api_key = req.openai_api_key or os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        raise HTTPException(status_code=400, detail="OPENAI_API_KEY not configured")

    event_queue: asyncio.Queue = asyncio.Queue()

    async def event_callback(event_type: str, data: dict):
        await event_queue.put({"type": event_type, **data})

    async def run():
        try:
            result = await run_autonomous(
                goal=req.goal,
                registered_agents=req.agents,
                api_key=api_key,
                event_callback=event_callback,
            )
            await event_queue.put({
                "type":   "autonomous_complete",
                "output": result.get("output", ""),
                "token_usage": result.get("token_usage", {}),
            })
        except Exception as e:
            await event_queue.put({"type": "autonomous_error", "error": str(e)})
        finally:
            await event_queue.put(None)

    async def event_generator():
        task = asyncio.create_task(run())
        while True:
            try:
                event = await asyncio.wait_for(event_queue.get(), timeout=300)
                if event is None:
                    break
                yield f"data: {json.dumps(event)}\n\n"
            except asyncio.TimeoutError:
                yield f"data: {json.dumps({'type': 'keepalive'})}\n\n"
        await task

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


@app.get("/tools")
async def list_tools():
    """List all available tools."""
    from runtime.tools import AVAILABLE_TOOLS
    return {
        "tools": [
            {"name": name, "description": tool.description}
            for name, tool in AVAILABLE_TOOLS.items()
        ]
    }


@app.get("/download")
async def download_file(path: str):
    """Serve a filled/generated file from workspace/uploads (security-scoped)."""
    import mimetypes

    def _normalize_download_path(raw_path: str) -> Path:
        value = unquote((raw_path or "").strip()).replace("\\", "/")
        if value.startswith("file://"):
            value = value[7:]
        if len(value) >= 4 and value[0] == "/" and value[2] == ":" and value[3] == "/":
            value = value[1:]

        candidates = [Path(value)]
        marker = "/workspace/uploads/"
        if marker in value:
            candidates.append(UPLOAD_DIR / value.split(marker, 1)[1])
        if value.startswith("workspace/uploads/"):
            candidates.append(UPLOAD_DIR / value.removeprefix("workspace/uploads/"))
        if value.startswith("/workspace/uploads/"):
            candidates.append(UPLOAD_DIR / value.removeprefix("/workspace/uploads/"))
        if value:
            candidates.append(UPLOAD_DIR / Path(value).name)

        for candidate in candidates:
            if candidate.exists():
                return candidate.resolve()
        return candidates[0].resolve()

    try:
        requested = _normalize_download_path(path)
        allowed   = UPLOAD_DIR.resolve()
        requested.relative_to(allowed)   # raises ValueError if outside uploads
    except ValueError:
        raise HTTPException(status_code=403, detail="Access denied: file must be in uploads directory")
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid path")

    if not requested.exists():
        raise HTTPException(status_code=404, detail="File not found")

    mime = mimetypes.guess_type(str(requested))[0] or "application/octet-stream"

    def _iter():
        with open(requested, "rb") as f:
            while chunk := f.read(65536):
                yield chunk

    return StreamingResponse(
        _iter(), media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{requested.name}"'},
    )


@app.post("/upload/pdf")
async def upload_pdf(file: UploadFile = File(...)):
    """Save an uploaded PDF to the workspace and return its path."""
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted.")
    dest = UPLOAD_DIR / file.filename
    with open(dest, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return {"path": str(dest), "filename": file.filename}


class PDFDetectRequest(BaseModel):
    path: str


@app.post("/pdf/detect-fields")
async def detect_pdf_fields(req: PDFDetectRequest):
    """
    Detect all form field labels in a PDF.

    Detection order:
      1. OCR.space API → GPT-4o-mini parse (PRIMARY — always runs first).
      2. Local geometry script (detect_form_fields.py) — fallback.

    After detection, saves a cache file  {pdf}.yuno_fields.json  that stores
    labels WITH fill-coordinates.  fill_pdf_form reads this cache so that
    both steps always use identical labels — no re-mapping needed.

    Returns:
      { fields: [{label, type, max_chars}], checkboxes: [{label, options:[{text}]}] }
    """
    import re as _re
    import httpx

    pdf_path = Path(req.path)
    if not pdf_path.exists():
        raise HTTPException(status_code=404, detail=f"PDF not found: {req.path}")

    ocr_api_key = os.getenv("OCR_SPACE_API_KEY", "")
    openai_key  = os.getenv("OPENAI_API_KEY", "")

    # ── helpers ───────────────────────────────────────────────────────────────
    def _norm(t: str) -> str:
        return _re.sub(r"[^a-z0-9]", "", t.lower())

    def _label_bbox(label: str, word_list: list) -> dict | None:
        """
        Find the bounding box of `label` text in a word list.

        Works with both fitz_words (native, exact) and ocr_words (scaled).
        Each item must have: text, left_pt, top_pt, right_pt, bottom_pt.

        Sliding window: concatenates normalized tokens until the full normalized
        label is contained.  Score = label_len / window_len (1.0 = exact).
        We ONLY fire on `nl in running` — never `running in nl` — to prevent
        a 1-char OCR word from matching every label that contains that character.
        """
        nl = _norm(label)
        if not nl:
            return None
        nw = [_norm(w["text"]) for w in word_list]
        best_score, best = 0.0, None
        for i in range(len(word_list)):
            running = ""
            for j in range(i, min(i + 10, len(word_list))):
                running += nw[j]
                if nl in running:
                    score = len(nl) / len(running)   # 1.0 for exact match
                    if score > best_score:
                        best_score = score
                        best = {
                            "right_pt":  max(word_list[k]["right_pt"]  for k in range(i, j + 1)),
                            "top_pt":    min(word_list[k]["top_pt"]    for k in range(i, j + 1)),
                            "bottom_pt": max(word_list[k]["bottom_pt"] for k in range(i, j + 1)),
                            "left_pt":   min(word_list[k]["left_pt"]   for k in range(i, j + 1)),
                        }
                    break   # extending only lowers the score
        # Embed page info from the first word in the match window so callers
        # know which page this label was found on (critical for multi-page PDFs).
        if best is not None:
            best["_page"]   = word_list[i].get("_page", 0)
            best["_page_h"] = word_list[i].get("_page_h", pdf_page_h)
        return best

    # ── get PDF page dimensions + native text/drawing extraction ─────────────
    # Processes ALL pages so multi-page PDFs get correct page indices.
    pdf_page_w, pdf_page_h = 595.0, 842.0

    # Per-page fitz data: page_idx → {words, text_boxes, cb_boxes, pw, ph}
    # Used for page-aware coordinate matching so fields land on the right page.
    _fitz_page_data: dict[int, dict] = {}

    # Flat combined lists across all pages (needed for global coord_words fallback)
    fitz_words:      list[dict] = []
    fitz_text_boxes: list[dict] = []
    fitz_cb_boxes:   list[dict] = []

    try:
        import fitz as _fitz
        _d = _fitz.open(str(pdf_path.resolve()))

        for _pi in range(_d.page_count):
            _p  = _d[_pi]
            _pw = _p.rect.width
            _ph = _p.rect.height

            if _pi == 0:
                pdf_page_w = _pw
                pdf_page_h = _ph

            _pg_words:      list[dict] = []
            _pg_text_boxes: list[dict] = []
            _pg_cb_boxes:   list[dict] = []

            # ── Native text words — tagged with their page index ──────────────
            for _w in _p.get_text("words"):
                _word_text = (_w[4] or "").strip()
                if _word_text:
                    _pg_words.append({
                        "text":      _word_text,
                        "left_pt":   float(_w[0]),
                        "top_pt":    float(_w[1]),
                        "right_pt":  float(_w[2]),
                        "bottom_pt": float(_w[3]),
                        "_page":     _pi,   # ← page index for later lookup
                        "_page_h":   _ph,
                    })

            # ── Drawn rectangles → form field boxes / checkboxes ──────────────
            for _drw in _p.get_drawings():
                _rect = _drw.get("rect")
                if _rect is None or _rect.is_empty:
                    continue
                _rw = float(_rect.width)
                _rh = float(_rect.height)
                if _rw <= 0 or _rh <= 0:
                    continue
                if 5 <= _rw <= 28 and 5 <= _rh <= 28 and abs(_rw - _rh) <= 5:
                    _pg_cb_boxes.append({
                        "x0": float(_rect.x0), "y0": float(_rect.y0),
                        "x1": float(_rect.x1), "y1": float(_rect.y1),
                        "w": _rw, "h": _rh,
                        "cx": (_rect.x0 + _rect.x1) / 2.0,
                        "cy": (_rect.y0 + _rect.y1) / 2.0,
                    })
                elif _rw >= 40 and 8 <= _rh <= 50:
                    _pg_text_boxes.append({
                        "x0": float(_rect.x0), "y0": float(_rect.y0),
                        "x1": float(_rect.x1), "y1": float(_rect.y1),
                        "w": _rw, "h": _rh,
                        "fill_x": float(_rect.x0) + 4.0,
                        "fill_y": float(_rect.y0) + _rh * 0.72,
                        "cx": (_rect.x0 + _rect.x1) / 2.0,
                        "cy": (_rect.y0 + _rect.y1) / 2.0,
                    })

            # Store per-page (dedup happens below after _dedup_rects is defined)
            _fitz_page_data[_pi] = {
                "words":      _pg_words,
                "text_boxes": _pg_text_boxes,   # deduped below
                "cb_boxes":   _pg_cb_boxes,     # deduped below
                "pw":         _pw,
                "ph":         _ph,
            }
            fitz_words.extend(_pg_words)
            fitz_text_boxes.extend(_pg_text_boxes)
            fitz_cb_boxes.extend(_pg_cb_boxes)

        _d.close()
    except Exception:
        pass

    # ── Deduplicate drawn boxes (Canva PDFs layer the same rect 3–4 times) ─────
    def _dedup_rects(boxes: list[dict]) -> list[dict]:
        """
        Keep only one representative rect per unique position.
        Prefer the orange-stroked one (the visible border).  Two rects are
        considered duplicates when their x0, y0, w, h all agree within 2 pt.
        """
        seen: list[dict] = []
        orange = (0.686, 0.298, 0.059)
        def _is_orange(c) -> bool:
            if not c:
                return False
            return abs(c[0]-orange[0])<0.05 and abs(c[1]-orange[1])<0.05 and abs(c[2]-orange[2])<0.05
        for box in boxes:
            duplicate = False
            for s in seen:
                if (abs(box["x0"]-s["x0"]) < 2 and abs(box["y0"]-s["y0"]) < 2
                        and abs(box["w"]-s["w"]) < 3 and abs(box["h"]-s["h"]) < 3):
                    duplicate = True
                    # Prefer the orange-bordered one
                    if _is_orange(box.get("color")):
                        seen[seen.index(s)] = box
                    break
            if not duplicate:
                seen.append(box)
        return seen

    fitz_text_boxes = _dedup_rects(fitz_text_boxes)
    fitz_cb_boxes   = _dedup_rects(fitz_cb_boxes)

    # Dedup per-page boxes too (used for page-aware nearest-box lookup)
    for _pi, _pgd in _fitz_page_data.items():
        _pgd["text_boxes"] = _dedup_rects(_pgd["text_boxes"])
        _pgd["cb_boxes"]   = _dedup_rects(_pgd["cb_boxes"])

    # ── Helpers: match labels/options to drawn boxes ──────────────────────────
    def _nearest_text_box(label_bb: dict, boxes: list | None = None) -> dict | None:
        """
        Find the drawn text-input rect closest to the RIGHT of (or BELOW) the label.
        Pass the page-specific box list so cross-page matches are impossible.
        Falls back to the global fitz_text_boxes when no per-page list is given.
        """
        search = boxes if boxes is not None else fitz_text_boxes
        if not label_bb or not search:
            return None
        lright = label_bb["right_pt"]
        ltop   = label_bb["top_pt"]
        lbot   = label_bb["bottom_pt"]
        lcy    = (ltop + lbot) / 2.0

        best_rect, best_dist = None, float("inf")
        for tb in search:
            if abs(tb["cy"] - lcy) <= 20 and tb["x0"] >= lright - 5:
                dist = (tb["x0"] - lright) + abs(tb["cy"] - lcy) * 1.5
                if dist < best_dist:
                    best_dist, best_rect = dist, tb
            elif tb["y0"] >= lbot and tb["cy"] <= lbot + 45:
                overlap = min(tb["x1"], label_bb["right_pt"]) - max(tb["x0"], label_bb["left_pt"])
                if overlap > 0:
                    dist = (tb["y0"] - lbot) + 200
                    if dist < best_dist:
                        best_dist, best_rect = dist, tb
        return best_rect if best_dist < 350 else None

    def _nearest_cb_box(opt_bb: dict, boxes: list | None = None) -> dict | None:
        """
        Find the checkbox square closest to the LEFT of the option text.
        Pass the page-specific box list to avoid cross-page matches.
        """
        search = boxes if boxes is not None else fitz_cb_boxes
        if not opt_bb or not search:
            return None
        oleft = opt_bb["left_pt"]
        ocy   = (opt_bb["top_pt"] + opt_bb["bottom_pt"]) / 2.0

        best_cb, best_dist = None, float("inf")
        for cb in search:
            if cb["cx"] > oleft + 10:
                continue
            if abs(cb["cy"] - ocy) > 20:
                continue
            dist = (oleft - cb["cx"]) + abs(cb["cy"] - ocy) * 2.0
            if dist < best_dist:
                best_dist, best_cb = dist, cb
        return best_cb if best_dist < 120 else None

    # ── Step 1: OCR.space + GPT-4o-mini (PRIMARY) ─────────────────────────────
    # Two calls:
    #   A) OCREngine 2  (no overlay) — best text quality for GPT label extraction
    #   B) OCREngine 1  (with overlay) — word bboxes (used ONLY when native PyMuPDF
    #      text extraction returned nothing, i.e. image-based/scanned PDFs)
    #
    # For COORDINATE matching we prefer fitz_words (exact) over ocr_words (scaled).
    raw_text  = ""
    ocr_words: list[dict] = []   # word bounding boxes in PDF points (OCR fallback)

    if ocr_api_key:
        # ── Call A: Engine 2, labels ──────────────────────────────────────────
        try:
            pdf_bytes = pdf_path.read_bytes()
            async with httpx.AsyncClient(timeout=60) as client:
                resp_a = await client.post(
                    "https://api.ocr.space/parse/image",
                    files={"file": (pdf_path.name, pdf_bytes, "application/pdf")},
                    data={
                        "apikey":            ocr_api_key,
                        "language":          "eng",
                        "OCREngine":         "2",
                        "isOverlayRequired": "false",
                        "detectOrientation": "true",
                        "scale":             "true",
                        "isTable":           "true",
                    },
                )
            for parsed in resp_a.json().get("ParsedResults", []):
                raw_text += parsed.get("ParsedText", "") + "\n"
        except Exception:
            raw_text = ""

        # ── Call B: Engine 1, coordinates ────────────────────────────────────
        try:
            async with httpx.AsyncClient(timeout=60) as client:
                resp_b = await client.post(
                    "https://api.ocr.space/parse/image",
                    files={"file": (pdf_path.name, pdf_bytes, "application/pdf")},
                    data={
                        "apikey":            ocr_api_key,
                        "language":          "eng",
                        "OCREngine":         "1",      # Engine 1 = full overlay support
                        "isOverlayRequired": "true",
                        "detectOrientation": "true",
                        "scale":             "true",
                    },
                )
            # Enumerate ParsedResults — one entry per PDF page
            for _ocr_pi, parsed in enumerate(resp_b.json().get("ParsedResults", [])):
                overlay = parsed.get("TextOverlay", {})
                # Use actual page dimensions from fitz if available (more accurate)
                _ocr_pw = _fitz_page_data.get(_ocr_pi, {}).get("pw", pdf_page_w)
                _ocr_ph = _fitz_page_data.get(_ocr_pi, {}).get("ph", pdf_page_h)
                img_w   = overlay.get("OriginalPageWidth",  1240) or 1240
                img_h   = overlay.get("OriginalPageHeight", 1754) or 1754
                sx      = _ocr_pw / img_w
                sy      = _ocr_ph / img_h
                for line in overlay.get("Lines", []):
                    for word in line.get("Words", []):
                        wt = (word.get("WordText") or "").strip()
                        if not wt:
                            continue
                        wl  = float(word.get("Left",   0))
                        wtp = float(word.get("Top",    0))
                        ww  = float(word.get("Width",  30))
                        wh  = float(word.get("Height", 12))
                        ocr_words.append({
                            "text":      wt,
                            "left_pt":   wl         * sx,
                            "top_pt":    wtp        * sy,
                            "right_pt":  (wl + ww)  * sx,
                            "bottom_pt": (wtp + wh) * sy,
                            "_page":     _ocr_pi,   # ← page index
                            "_page_h":   _ocr_ph,
                        })
        except Exception:
            pass   # coordinates unavailable — fields will be cached without coords

    fields:     list[dict] = []
    checkboxes: list[dict] = []

    # Choose the best word list for label matching:
    # fitz_words (native, exact) >> ocr_words (OCR-scaled, approximate)
    coord_words = fitz_words if fitz_words else ocr_words

    # For the GPT prompt use OCR text (Engine 2) if available, else fitz plain text
    gpt_text = raw_text.strip()
    if not gpt_text and fitz_words:
        gpt_text = " ".join(w["text"] for w in fitz_words)

    if gpt_text and openai_key:
        try:
            from langchain_openai import ChatOpenAI
            model  = ChatOpenAI(model="gpt-4o-mini", api_key=openai_key, temperature=0)
            prompt = f"""You are analyzing OCR text extracted from a form PDF.
Identify EVERY field where a person writes or selects information.

OUTPUT FORMAT — one entry per line, nothing else:

For free-text fields (name, date, address, etc.):
  <Field Label>

For selection / checkbox / radio groups where the user picks ONE option:
  CHECKBOX: <Group Label> | <Option1> | <Option2> | ...

IMPORTANT — mark as CHECKBOX if the form has:
  • Male / Female (or M / F)
  • Yes / No
  • Multiple membership tiers  (e.g. Basic / Standard / Premium)
  • Duration options           (e.g. 6 Months / 12 Months / 24 Months)
  • Status options             (e.g. Active / Inactive / Pending)
  • Any list of 2–6 mutually exclusive tick-box / radio-button options

Rules:
- No headings, instructions, logos, footers, or page numbers.
- No duplicates. No numbering. No explanations.
- Keep labels short (1–6 words).

OCR Text:
{gpt_text[:3500]}"""

            response = model.invoke([{"role": "user", "content": prompt}])
            seen: set[str] = set()

            for line in response.content.strip().split("\n"):
                line = line.strip(" -•*")
                if not line:
                    continue
                if line.upper().startswith("CHECKBOX:"):
                    parts = [p.strip() for p in line[9:].split("|") if p.strip()]
                    if len(parts) >= 2:
                        label   = parts[0]
                        options = [{"text": o} for o in parts[1:]]
                        if label and label not in seen:
                            seen.add(label)
                            opts_with_coords = []
                            cb_page    = 0
                            cb_page_h  = pdf_page_h
                            for opt in options:
                                bb = _label_bbox(opt["text"], coord_words)
                                if bb:
                                    # Determine which page this option is on
                                    opt_page   = bb.get("_page", 0)
                                    opt_page_h = bb.get("_page_h", pdf_page_h)
                                    cb_page    = opt_page
                                    cb_page_h  = opt_page_h
                                    oh = bb["bottom_pt"] - bb["top_pt"]
                                    bs = max(oh, 8.0)
                                    # Use page-specific cb_boxes to avoid cross-page matches
                                    pg_cb_boxes = _fitz_page_data.get(opt_page, {}).get("cb_boxes", fitz_cb_boxes)
                                    drw_cb = _nearest_cb_box(bb, pg_cb_boxes)
                                    if drw_cb:
                                        opts_with_coords.append({
                                            "text": opt["text"],
                                            "x":    drw_cb["x0"],
                                            "y":    drw_cb["y0"],
                                            "w":    drw_cb["w"],
                                            "h":    drw_cb["h"],
                                            "ocr_flat": True,
                                        })
                                    else:
                                        opts_with_coords.append({
                                            "text": opt["text"],
                                            "x":    max(0.0, bb["left_pt"] - bs - 3),
                                            "y":    bb["top_pt"],
                                            "w":    bs, "h": bs,
                                            "ocr_flat": True,
                                        })
                                else:
                                    opts_with_coords.append({"text": opt["text"]})
                            checkboxes.append({
                                "label":       label,
                                "options":     opts_with_coords,
                                "page":        cb_page,    # ← correct page
                                "page_height": cb_page_h,
                            })
                else:
                    if line and line not in seen and len(line) <= 60:
                        seen.add(line)
                        bb = _label_bbox(line, coord_words)
                        field_entry: dict = {
                            "label":    line,
                            "type":     "free_text",
                            "max_chars": 100,
                        }
                        if bb:
                            # Resolve which page this label belongs to
                            label_page   = bb.get("_page", 0)
                            label_page_h = bb.get("_page_h", pdf_page_h)
                            label_page_w = _fitz_page_data.get(label_page, {}).get("pw", pdf_page_w)
                            lh = bb["bottom_pt"] - bb["top_pt"]
                            # Use page-specific text boxes to avoid cross-page matches
                            pg_text_boxes = _fitz_page_data.get(label_page, {}).get("text_boxes", fitz_text_boxes)
                            drw_box = _nearest_text_box(bb, pg_text_boxes)
                            if drw_box:
                                field_entry.update({
                                    "x":           drw_box["fill_x"],
                                    "y":           drw_box["fill_y"],
                                    "line_end_x":  drw_box["x1"] - 4,
                                    "max_chars":   max(1, int((drw_box["w"] - 8) / 6.0)),
                                    "page":        label_page,    # ← correct page
                                    "page_height": label_page_h,
                                    "ocr_flat":    True,
                                })
                            else:
                                gap = 20.0
                                fx  = bb["right_pt"] + gap
                                fy  = bb["top_pt"] + lh * 0.75
                                field_entry.update({
                                    "x":           fx,
                                    "y":           fy,
                                    "line_end_x":  label_page_w - 10,
                                    "max_chars":   max(1, int((label_page_w - fx - 10) / 6.0)),
                                    "page":        label_page,    # ← correct page
                                    "page_height": label_page_h,
                                    "ocr_flat":    True,
                                })
                        fields.append(field_entry)
        except Exception:
            pass

    # ── Save cache & return if OCR+GPT succeeded ──────────────────────────────
    if fields or checkboxes:
        try:
            cache = {"fields": fields, "checkboxes": checkboxes, "source": "ocr"}
            Path(str(pdf_path.resolve()) + ".yuno_fields.json").write_text(
                json.dumps(cache), encoding="utf-8"
            )
        except Exception:
            pass
        # Return label-only response for the frontend (no coords needed in UI)
        ui_fields = [{"label": f["label"], "type": f.get("type","free_text"),
                      "max_chars": f.get("max_chars", 100)} for f in fields]
        ui_cbs    = [{"label": c["label"],
                      "options": [{"text": o["text"]} for o in c.get("options", [])]}
                     for c in checkboxes]
        return {"fields": ui_fields, "checkboxes": ui_cbs}

    # ── Step 2: Local geometry script (FALLBACK) ──────────────────────────────
    script = Path(__file__).parent / "scripts" / "detect_form_fields.py"
    if script.exists():
        try:
            result = subprocess.run(
                [sys.executable, str(script), str(pdf_path.resolve())],
                capture_output=True, text=True, timeout=120,
            )
            if result.returncode == 0 and result.stdout.strip():
                local_data = json.loads(result.stdout)
                if not local_data.get("error") and (
                    local_data.get("fields") or local_data.get("checkboxes")
                ):
                    # Cache geometry result too
                    try:
                        cache = {**local_data, "source": "geometry"}
                        Path(str(pdf_path.resolve()) + ".yuno_fields.json").write_text(
                            json.dumps(cache), encoding="utf-8"
                        )
                    except Exception:
                        pass
                    return local_data
        except Exception:
            pass

    return {"fields": [], "checkboxes": []}


@app.post("/pdf/debug-fields")
async def debug_pdf_fields(req: PDFDetectRequest):
    """
    Return the raw cached field data including fill coordinates.
    Useful for diagnosing why text lands in the wrong place.
    """
    cache_path = Path(req.path + ".yuno_fields.json")
    if not cache_path.exists():
        cache_path = Path(req.path).resolve()
        cache_path = Path(str(cache_path) + ".yuno_fields.json")
    if not cache_path.exists():
        return {"error": "No cache found — run /pdf/detect-fields first", "fields": [], "checkboxes": []}
    try:
        data = json.loads(cache_path.read_text(encoding="utf-8"))
        return {
            "source": data.get("source", "?"),
            "fields": [
                {
                    "label": f.get("label"),
                    "x": f.get("x"), "y": f.get("y"),
                    "line_end_x": f.get("line_end_x"),
                    "max_chars": f.get("max_chars"),
                    "has_coords": "x" in f,
                    "coord_source": "drawn_rect" if f.get("ocr_flat") and "line_end_x" in f else "estimated",
                }
                for f in data.get("fields", [])
            ],
            "checkboxes": [
                {
                    "label": c.get("label"),
                    "options": [
                        {"text": o.get("text"), "x": o.get("x"), "y": o.get("y"),
                         "w": o.get("w"), "h": o.get("h"), "has_coords": "x" in o}
                        for o in c.get("options", [])
                    ],
                }
                for c in data.get("checkboxes", [])
            ],
        }
    except Exception as e:
        return {"error": str(e)}


@app.post("/upload/files")
async def upload_files(files: list[UploadFile] = File(...)):
    """Save uploaded files or browser folder selections to workspace/uploads."""
    saved = []
    for file in files:
        raw_name = (file.filename or "upload").replace("\\", "/").lstrip("/")
        safe_parts = [p for p in raw_name.split("/") if p not in ("", ".", "..")]
        rel = Path(*safe_parts) if safe_parts else Path("upload")
        dest = UPLOAD_DIR / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        with open(dest, "wb") as f:
            shutil.copyfileobj(file.file, f)
        saved.append({"path": str(dest), "filename": raw_name, "size": dest.stat().st_size})
    return {"files": saved, "count": len(saved)}


class ExportRequest(BaseModel):
    content: str
    filename: str = "agent_response"
    format: str = "pdf"          # "pdf" or "excel"
    title: str = "Agent Response"


def _make_pdf(content: str, title: str) -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 not installed. Run: pip install fpdf2")

    class _PDF(FPDF):
        def header(self):
            self.set_font("Helvetica", "B", 13)
            self.cell(0, 10, title, align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_font("Helvetica", "", 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 6, "Generated by Yuno AI", align="C", new_x="LMARGIN", new_y="NEXT")
            self.set_text_color(0, 0, 0)
            self.ln(3)

        def footer(self):
            self.set_y(-13)
            self.set_font("Helvetica", "I", 8)
            self.set_text_color(150, 150, 150)
            self.cell(0, 10, f"Page {self.page_no()}", align="C")

    pdf = _PDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.set_font("Helvetica", "", 11)

    for line in content.split("\n"):
        clean = line.encode("latin-1", errors="replace").decode("latin-1")
        if clean.startswith("# "):
            pdf.set_font("Helvetica", "B", 13)
            pdf.multi_cell(0, 7, clean[2:])
            pdf.set_font("Helvetica", "", 11)
        elif clean.startswith("## "):
            pdf.set_font("Helvetica", "B", 11)
            pdf.multi_cell(0, 7, clean[3:])
            pdf.set_font("Helvetica", "", 11)
        elif clean.startswith("- ") or clean.startswith("* "):
            pdf.multi_cell(0, 6, f"  • {clean[2:]}")
        elif clean.startswith("---") or clean.startswith("==="):
            pdf.ln(2)
            pdf.line(pdf.get_x(), pdf.get_y(), pdf.get_x() + 180, pdf.get_y())
            pdf.ln(3)
        else:
            pdf.multi_cell(0, 6, clean if clean.strip() else " ")

    return bytes(pdf.output())


def _make_excel(content: str, title: str) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font  = Font(bold=True, color="FFFFFF", size=12)
    header_fill  = PatternFill("solid", fgColor="EC4899")
    bold_font    = Font(bold=True)
    wrap_align   = Alignment(wrap_text=True, vertical="top")

    row = 1
    for line in content.split("\n"):
        stripped = line.strip()
        if not stripped:
            row += 1
            continue

        if stripped.startswith("# ") or stripped.startswith("## "):
            text = stripped.lstrip("#").strip()
            cell = ws.cell(row=row, column=1, value=text)
            cell.font = Font(bold=True, size=13 if stripped.startswith("# ") else 11)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        elif "|" in stripped and not stripped.replace("|", "").replace("-", "").replace(" ", "") == "":
            parts = [p.strip() for p in stripped.split("|") if p.strip()]
            if not parts:
                row += 1
                continue
            is_header_sep = all(set(p) <= {"-", " ", ":"} for p in parts)
            if is_header_sep:
                row += 1
                continue
            for col, val in enumerate(parts, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.alignment = wrap_align
                # Style first row of each table as header
                prev = ws.cell(row=row - 1, column=1).value
                if prev is None or (row >= 2 and ws.cell(row=row-1, column=1).font.bold):
                    cell.font = header_font
                    cell.fill = header_fill
        else:
            cell = ws.cell(row=row, column=1, value=stripped)
            cell.alignment = wrap_align
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for c in col:
            try:
                max_len = max(max_len, len(str(c.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@app.post("/export")
async def export_content(req: ExportRequest):
    """Export agent response content as PDF or Excel."""
    fmt = req.format.lower()
    if fmt == "pdf":
        data = _make_pdf(req.content, req.title)
        return Response(
            content=data,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{req.filename}.pdf"'},
        )
    elif fmt in ("excel", "xlsx"):
        data = _make_excel(req.content, req.title)
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{req.filename}.xlsx"'},
        )
    else:
        raise HTTPException(status_code=400, detail="format must be 'pdf' or 'excel'")


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=False)
